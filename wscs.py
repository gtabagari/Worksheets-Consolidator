from openpyxl import load_workbook
from xlsxwriter import Workbook
from os import listdir, path, remove
import getopt
import sys


def copy_many_to_one(files_path, new_file):
    # new_file = 'collected.xlsx'
    global ws1
    if path.exists(new_file):
        try:
            remove(new_file)
        except:
            sys.exit('File not deleted.')
    if not path.exists(new_file):
        workbook = Workbook(new_file)
        print("The new workbook file was created.")
        workbook.close()

    source_file_list = [f for f in listdir(files_path) if f.endswith('.xlsx')]

    wb_target = load_workbook(filename=new_file)
    print('------------Begin collection----------')
    if source_file_list:
        for file in source_file_list:
            print('Filename: ', file)
            try:
                wb1 = load_workbook(filename=files_path + file)
                for count in range(0, len(wb1.sheetnames)):
                    ws1 = wb1.worksheets[count]
                    ws2 = wb_target.create_sheet((file + "_" + ws1.title)[:30])
                    print(ws1)
                    for row in ws1:
                        for cell in row:
                            ws2[cell.coordinate].value = cell.value
                    wb_target.save(new_file)
            except Exception as e:
                print("Something went wrong! (sheet name: " + ws1.title + ")")
    else:
        print('"xlsx" file cannot be found.')


def main(argv):
    inputfile_path = ''
    outputfile = ''
    try:
        opts, args = getopt.getopt(argv, "hi:o:", ["ifile=", "ofile="])
    except getopt.GetoptError:
        print('Usage: <exec> -i <input files path> -o <output file>')
        sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            print('Usage: <exec> -i <input files path> -o <output file>')
            sys.exit()
        elif opt in ("-i", "--ifile"):
            inputfile_path = arg
        elif opt in ("-o", "--ofile"):
            outputfile = arg
    print('Input filepath is "', inputfile_path)
    print('Output file is "', outputfile)

    if not inputfile_path or not outputfile:
        print('Usage: <exec> -i <input files path> -o <output file>')
        sys.exit()

    if inputfile_path[-1:] != '\/':
        inputfile_path += '\/'
    copy_many_to_one(inputfile_path, outputfile)


if __name__ == "__main__":
    main(sys.argv[1:])
